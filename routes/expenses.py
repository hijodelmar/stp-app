from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_from_directory, jsonify
from flask_login import login_required, current_user
from extensions import db
from models import Expense, Supplier
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from utils.auth import role_required

bp = Blueprint('expenses', __name__)

@bp.route('/stats-data')
@login_required
@role_required(['access_expenses'])
def stats_data():
    # Data for Pie Chart (By Category)
    category_stats = db.session.query(
        Expense.category, 
        db.func.sum(Expense.amount_ttc)
    ).group_by(Expense.category).all()
    
    categories = {
        'restaurant': 'Restaurant',
        'transport': 'Transport',
        'material': 'Matériel',
        'urssaf': 'Charges',
        'salary': 'Salaire',
        'other': 'Autre'
    }
    
    pie_labels = [categories.get(cat, cat) for cat, _ in category_stats]
    pie_data = [float(amount) for _, amount in category_stats]
    
    # Data for Bar Chart (By Month - Current Year)
    current_year = datetime.now().year
    monthly_stats = db.session.query(
        db.func.strftime('%m', Expense.date),
        db.func.sum(Expense.amount_ttc)
    ).filter(db.func.strftime('%Y', Expense.date) == str(current_year))\
    .group_by(db.func.strftime('%m', Expense.date)).all()
    
    # Initialize all months with 0
    bar_data = [0] * 12
    for month_str, amount in monthly_stats:
        month_idx = int(month_str) - 1
        bar_data[month_idx] = float(amount)
        
    bar_labels = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc']
    
    return jsonify({
        'pie': {'labels': pie_labels, 'data': pie_data},
        'bar': {'labels': bar_labels, 'data': bar_data}
    })

@bp.route('/export/excel')
@login_required
@role_required(['access_expenses'])
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file
    
    # Re-use the same filtering logic as index
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    
    query = Expense.query.filter_by(created_by_id=current_user.id)
    
    if search:
        query = query.filter(
            (Expense.description.ilike(f'%{search}%')) |
            (Expense.amount_ttc.like(f'%{search}%'))
        )
    if category:
        query = query.filter_by(category=category)
        
    expenses = query.order_by(Expense.date.desc()).all()
    
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dépenses"
    
    # Headers
    headers = ['Date', 'Description', 'Catégorie', 'Fournisseur', 'Montant TTC', 'TVA', 'Montant HT']
    ws.append(headers)
    
    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Add Data
    for expense in expenses:
        # Calculate HT/TVA if not present (simple estimation if needed, but model usually doesnt have it yet)
        # For now simply listing what we have. If we had HT/TVA fields we'd use them.
        # Assuming standard request was to just export what is in the list.
        supplier_name = expense.supplier.raison_sociale if expense.supplier else ''
        ws.append([
            expense.date.strftime('%d/%m/%Y'),
            expense.description,
            expense.category,
            supplier_name,
            expense.amount_ttc,
            expense.tva,
            expense.amount_ht
        ])
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Depenses_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/export/print')
@login_required
@role_required(['access_expenses'])
def print_view():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    
    query = Expense.query.filter_by(created_by_id=current_user.id)
    
    if search:
        query = query.filter(
            (Expense.description.ilike(f'%{search}%')) |
            (Expense.amount_ttc.like(f'%{search}%'))
        )
    if category:
        query = query.filter_by(category=category)
        
    expenses = query.order_by(Expense.date.desc()).all()
    total_ttc = sum(e.amount_ttc for e in expenses)
    total_ht = sum(e.amount_ht for e in expenses)
    total_tva = sum(e.tva for e in expenses)
    
    return render_template('expenses/print.html', expenses=expenses, 
                           total_ttc=total_ttc, 
                           total_ht=total_ht,
                           total_tva=total_tva,
                           now=datetime.now())

    return render_template('expenses/print.html', expenses=expenses, total_ttc=total_ttc, now=datetime.now())

@bp.route('/duplicate/<int:id>')
@login_required
@role_required(['access_expenses'])
def duplicate(id):
    original = Expense.query.get_or_404(id)
    if original.created_by_id != current_user.id:
        flash('Accès refusé.', 'danger')
        return redirect(url_for('expenses.index'))
    
    # Create copy
    new_expense = Expense(
        date=datetime.now().date(), # Reset date to today
        description=f"{original.description} (Copie)",
        category=original.category,
        amount_ttc=original.amount_ttc,
        supplier_id=original.supplier_id,
        created_by_id=current_user.id
        # Note: We do NOT copy attachments/proof_path as those are specific to the receipt
    )
    
    db.session.add(new_expense)
    db.session.commit()
    
    flash('Dépense dupliquée. Vous pouvez maintenant la modifier/valider.', 'success')
    return redirect(url_for('expenses.edit', id=new_expense.id))

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/receipts/<path:filename>')
@login_required
@role_required(['access_expenses'])
def get_receipt(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@bp.route('/')
@login_required
@role_required(['access_expenses'])
def index():
    # Filters
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
        
    expenses = Expense.query.filter(
        Expense.date >= start_date, 
        Expense.date < end_date
    ).order_by(Expense.date.desc()).all()
    
    total_ht = sum(e.amount_ht for e in expenses)
    total_ttc = sum(e.amount_ttc for e in expenses)
    to_reimburse = sum(e.amount_ttc for e in expenses if e.payment_method == 'personal_funds' and not e.is_reimbursed)
    
    return render_template('expenses/index.html', 
                         expenses=expenses, 
                         total_ht=total_ht, 
                         total_ttc=total_ttc,
                         to_reimburse=to_reimburse,
                         current_month=month,
                         current_year=year,
                         now=datetime.now())

@bp.route('/add', methods=['GET', 'POST'])
@login_required
@role_required(['access_expenses'])
def add():
    suppliers = Supplier.query.order_by(Supplier.raison_sociale).all()
    
    if request.method == 'POST':
        try:
            # Initial Data from Form
            date_str = request.form.get('date')
            description = request.form.get('description')
            amount_input = request.form.get('amount_ttc')
            tva_input = request.form.get('tva')
            category = request.form.get('category')
            payment_method = request.form.get('payment_method')
            supplier_id = request.form.get('supplier_id')
            
            # Files Processing & OCR
            ocr_data = {}
            saved_attachments = [] # List of tuples (abs_path, rel_path, filename)
            
            files = request.files.getlist('proof')
            if files and files[0].filename != '':
                # Save the first file to run OCR on it
                first_file = files[0]
                if allowed_file(first_file.filename):
                    # Use temp description if missing
                    safe_desc = secure_filename(description[:30]) if description else "NewExpense"
                    safe_date = date_str if date_str else datetime.now().strftime('%Y-%m-%d')
                    
                    filename = secure_filename(f"{safe_date}_{safe_desc}_{first_file.filename}")
                    rel_path = os.path.join('expenses', str(datetime.now().year), str(datetime.now().month))
                    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], rel_path)
                    os.makedirs(abs_path, exist_ok=True)
                    
                    save_path_abs = os.path.join(abs_path, filename)
                    save_path_rel = os.path.join(rel_path, filename).replace('\\', '/')
                    
                    first_file.save(save_path_abs)
                    
                    # Store for later linking
                    saved_attachments.append((save_path_abs, save_path_rel, first_file.filename))
                    
                    # Run OCR
                    try:
                        from utils.ocr import extract_expense_data
                        ocr_data = extract_expense_data(save_path_abs)
                    except Exception as e:
                        print(f"OCR Error: {e}")
            
            # Apply OCR Data if form fields are empty
            if not date_str and ocr_data.get('date'):
                try:
                    # OCR returns DD/MM/YYYY, we need YYYY-MM-DD
                    d_obj = datetime.strptime(ocr_data['date'], '%d/%m/%Y')
                    date_str = d_obj.strftime('%Y-%m-%d')
                except:
                    pass # Keep None/Today
            
            if (not amount_input or float(amount_input) == 0) and ocr_data.get('amount_ttc'):
                amount_input = ocr_data['amount_ttc']
                
            if (not tva_input or float(tva_input) == 0) and ocr_data.get('tva'):
                tva_input = ocr_data['tva']
                
            if (not category or category == 'other') and ocr_data.get('category'):
                category = ocr_data['category']

            # Defaults if still empty
            if not date_str:
                date_str = datetime.now().strftime('%Y-%m-%d')
            if not description:
                description = "Dépense (Scan)" if ocr_data else "Nouvelle Dépense"
            
            amount_ttc = float(amount_input) if amount_input else 0.0
            tva = float(tva_input) if tva_input else 0.0
            amount_ht = amount_ttc - tva
            
            # Create Expense Object
            expense = Expense(
                date=datetime.strptime(date_str, '%Y-%m-%d'),
                description=description,
                amount_ht=amount_ht,
                tva=tva,
                amount_ttc=amount_ttc,
                category=category or 'other',
                payment_method=payment_method or 'personal_funds',
                supplier_id=supplier_id if supplier_id else None,
                proof_path=None, # Will be set below
                created_by_id=current_user.id
            )
            
            db.session.add(expense)
            db.session.flush() # Get ID
            
            # Link Attachments (First one checks out)
            from models import ExpenseAttachment
            
            # 1. Link the first file (already saved)
            if saved_attachments:
                path_abs, path_rel, fname = saved_attachments[0]
                att = ExpenseAttachment(expense_id=expense.id, file_path=path_rel, filename=fname)
                db.session.add(att)
                expense.proof_path = path_rel # Legacy support
            
            # 2. Save remaining files (skipping the first one)
            for i, file in enumerate(files):
                if i == 0: continue # Already handled
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(f"{date_str}_{description[:30]}_{file.filename}")
                    rel_path = os.path.join('expenses', str(datetime.now().year), str(datetime.now().month))
                    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], rel_path)
                    os.makedirs(abs_path, exist_ok=True)
                    
                    file.save(os.path.join(abs_path, filename))
                    saved_rel_path = os.path.join(rel_path, filename).replace('\\', '/')
                    
                    attachment = ExpenseAttachment(
                        expense_id=expense.id,
                        file_path=saved_rel_path,
                        filename=file.filename
                    )
                    db.session.add(attachment)

            db.session.commit()
            
            msg = 'Enregistrement de la dépense effectuée.'
            if ocr_data:
                msg += ' (OCR : Données détectées)'
            flash(msg, 'success')
            
            return redirect(url_for('expenses.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
            
    return render_template('expenses/form.html', suppliers=suppliers, now=datetime.now())

@bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['access_expenses'])
def edit(id):
    expense = Expense.query.get_or_404(id)
    suppliers = Supplier.query.order_by(Supplier.raison_sociale).all()
    
    if request.method == 'POST':
        try:
            expense.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
            expense.description = request.form.get('description')
            expense.amount_ttc = float(request.form.get('amount_ttc'))
            expense.tva = float(request.form.get('tva', 0))
            expense.amount_ht = expense.amount_ttc - expense.tva
            expense.category = request.form.get('category')
            expense.payment_method = request.form.get('payment_method')
            expense.supplier_id = request.form.get('supplier_id') or None
            
            # Multiple Files Upload (Append)
            files = request.files.getlist('proof')
            for file in files:
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(f"{request.form.get('date')}_{expense.description[:30]}_{file.filename}")
                    rel_path = os.path.join('expenses', str(datetime.now().year), str(datetime.now().month))
                    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], rel_path)
                    os.makedirs(abs_path, exist_ok=True)
                    
                    file.save(os.path.join(abs_path, filename))
                    saved_rel_path = os.path.join(rel_path, filename).replace('\\', '/')
                    
                    # Create Attachment Record
                    from models import ExpenseAttachment
                    attachment = ExpenseAttachment(
                        expense_id=expense.id,
                        file_path=saved_rel_path,
                        filename=file.filename
                    )
                    db.session.add(attachment)
                    
                    if not expense.proof_path:
                        expense.proof_path = saved_rel_path
            
            # Update history
            expense.updated_by_id = current_user.id
            expense.updated_at = datetime.utcnow()
            
            
            db.session.commit()
            flash('Enregistrement de la dépense effectuée.', 'success')
            return redirect(url_for('expenses.index'))
        except Exception as e:
            flash(f'Erreur: {str(e)}', 'danger')

    return render_template('expenses/form.html', expense=expense, suppliers=suppliers)

@bp.route('/delete_attachment/<int:id>', methods=['POST'])
@login_required
@role_required(['access_expenses'])
def delete_attachment(id):
    from models import ExpenseAttachment
    attachment = ExpenseAttachment.query.get_or_404(id)
    expense = attachment.expense
    
    try:
        # Remove file from disk
        abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], attachment.file_path)
        if os.path.exists(abs_path):
            os.remove(abs_path)
            
        # Remove from DB
        db.session.delete(attachment)
        
        # If this was the legacy proof_path, try to set another one or clear it
        if expense.proof_path == attachment.file_path:
            other_attachment = ExpenseAttachment.query.filter(
                ExpenseAttachment.expense_id == expense.id,
                ExpenseAttachment.id != attachment.id
            ).first()
            expense.proof_path = other_attachment.file_path if other_attachment else None
            
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/delete/<int:id>', methods=['POST'])
@login_required
@role_required(['access_expenses'])
def delete(id):
    expense = Expense.query.get_or_404(id)
    
    # Delete all attachments from disk
    for attachment in expense.attachments:
        abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], attachment.file_path)
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except:
                pass

    # Legacy cleanup
    if expense.proof_path:
         abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], expense.proof_path)
         if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except:
                pass
                
    db.session.delete(expense)
    db.session.commit()
    flash('Dépense supprimée.', 'success')
    return redirect(url_for('expenses.index'))

@bp.route('/scan', methods=['POST'])
@login_required
@role_required(['access_expenses'])
def scan_receipt():
    print("DEBUG: /scan route hit")
    if 'proof' not in request.files:
        print("DEBUG: No proof file in request")
        return jsonify({'success': False, 'error': 'Aucun fichier reçu'}), 400
        
    file = request.files['proof']
    print(f"DEBUG: File received: {file.filename}")
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Nom de fichier vide'}), 400
        
    if file and allowed_file(file.filename):
        try:
            # Save temporarily
            filename = secure_filename(f"temp_scan_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp', filename)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            file.save(abs_path)
            print(f"DEBUG: File saved to {abs_path}")
            
            # Run OCR
            print("DEBUG: Calling extract_expense_data...")
            from utils.ocr import extract_expense_data
            data = extract_expense_data(abs_path)
            print(f"DEBUG: OCR Data result: {data}")
            
            # Clean up temp file
            try:
                os.remove(abs_path)
            except:
                pass
                
            # Convert date format dd/mm/yyyy to yyyy-mm-dd for input[type=date]
            if data.get('date'):
                try:
                    d_obj = datetime.strptime(data['date'], '%d/%m/%Y')
                    data['date'] = d_obj.strftime('%Y-%m-%d')
                except:
                    data['date'] = None
            
            return jsonify({'success': True, 'data': data})
            
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
            
    return jsonify({'success': False, 'error': 'Fichier invalide'}), 400
